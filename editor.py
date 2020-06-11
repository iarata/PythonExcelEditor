'''
Date: June 11, 2020
Reading all the data in the Excel file finding mean and std of each column and writing it down.
And for CorrectedData sheet, comparing values with the number that we got from "MEAN-2STD" and
if the numebr is bigger or lower than that number we remove it and calculate it again.
@author: Hajebrahimi Alireza - Ritsumeikan University
'''
import os
from statistics import stdev

try:
    import openpyxl
except ImportError:
    print("Trying to Install required module: openpyxl\n")
    os.system('python -m pip install openpyxl')
# -- above lines try to install openpyxl module if not present
# -- if all went well, import required module again (for global access)
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font

# Defining the Path to the File and Loading the current Excel File.
workingPath = str(os.path.dirname(os.path.abspath(__file__)))
myPath = str(input(
    '\nFirst, Put this python file and the excel file that you want to change.\nPlease write Excel file name with extenstion: '))
wook = openpyxl.load_workbook(f'{workingPath}/{myPath}')
CorrectedDataSheet = wook['CorrectedData']
rawDataSheet = wook['RawData']

rawDataMean = {}
rawDataSTD = {}


def Average(lst):
    # The Average Function
    return sum(lst)/len(lst)


def STDCal(lst):
    # Calculating the Standard Deviation
    return stdev(lst)


# Finding the mean and STD for RawData
def RawDataSetup(rawSTD, rawMean):

    # [ RawData ]: Going throw all the columns and for the rows from  2~19.
    for col in rawDataSheet.iter_cols(min_row=2, max_row=19, min_col=3, max_col=rawDataSheet.max_column):

        # Stroing all the data of the column in a list.
        currentColCellsValue = []
        for cell in col:
            if cell.value != None:
                currentColCellsValue.append(cell.value)
        # print(f"{get_column_letter(cell.column)}22")
        # rawDataSheet[f'{get_column_letter(cell.column)}22'] = Average(
        #     currentColCellsValue)
        # rawDataSheet[f'{get_column_letter(cell.column)}23'] = STDCal(
        #     currentColCellsValue)
        rawDataSheet[f'{get_column_letter(cell.column)}22'] = f'=AVERAGE({get_column_letter(cell.column)}2:{get_column_letter(cell.column)}19)'
        rawDataSheet[f'{get_column_letter(cell.column)}23'] = f'=STDEV({get_column_letter(cell.column)}2:{get_column_letter(cell.column)}19)'

        rawSTD[f'{get_column_letter(cell.column)}'] = STDCal(
            currentColCellsValue)
        rawMean[f'{get_column_letter(cell.column)}'] = Average(
            currentColCellsValue)

        rawDataSheet[f'{get_column_letter(cell.column)}22'].font = Font(
            name='Consolas', sz=16, italic=True)
        rawDataSheet[f'{get_column_letter(cell.column)}23'].font = Font(
            name='Consolas', sz=16, italic=True)


def CorrectedDataSetup(rawSTD, rawMean):
    # [ CorrectedData ]: Going throw all the columns and rows and comparing them with MEAN-+2STD of RawData
    # and then removing them from the File and at the end calculating the new MEAN and STD
    howManySTD = int(input('How Many STD? (Only Numbers: 1, 2, 3): '))

    for col in CorrectedDataSheet.iter_cols(min_row=2, max_row=19, min_col=3, max_col=CorrectedDataSheet.max_column):

        correctedDatas = []

        for cell in col:
            if cell.value != None and (cell.value <= (rawMean[f'{get_column_letter(cell.column)}']-(howManySTD*rawSTD[f'{get_column_letter(cell.column)}'])) or cell.value >= (rawMean[f'{get_column_letter(cell.column)}']+(howManySTD*rawSTD[f'{get_column_letter(cell.column)}']))):
                CorrectedDataSheet[f'{get_column_letter(cell.column)}{cell.row}'] = None
            elif cell.value != None:
                correctedDatas.append(cell.value)

        CorrectedDataSheet[f'{get_column_letter(cell.column)}22'] = f'=AVERAGE({get_column_letter(cell.column)}2:{get_column_letter(cell.column)}19)'
        CorrectedDataSheet[f'{get_column_letter(cell.column)}23'] = f'=STDEV({get_column_letter(cell.column)}2:{get_column_letter(cell.column)}19)'
        CorrectedDataSheet[f'{get_column_letter(cell.column)}22'].font = Font(
            name='Consolas', sz=16, italic=True)
        CorrectedDataSheet[f'{get_column_letter(cell.column)}23'].font = Font(
            name='Consolas', sz=16, italic=True)

        CorrectedDataSheet[f'{get_column_letter(cell.column)}25'] = rawMean[f'{get_column_letter(cell.column)}']-(
            howManySTD*rawSTD[f'{get_column_letter(cell.column)}'])
        CorrectedDataSheet[f'{get_column_letter(cell.column)}24'] = rawMean[f'{get_column_letter(cell.column)}']+(
            howManySTD*rawSTD[f'{get_column_letter(cell.column)}'])
        CorrectedDataSheet[f'{get_column_letter(cell.column)}25'].font = Font(
            name='Consolas', sz=16, italic=True)
        CorrectedDataSheet[f'{get_column_letter(cell.column)}24'].font = Font(
            name='Consolas', sz=16, italic=True)

    CorrectedDataSheet['B25'] = f'MEAN-{howManySTD}STD (RawData)'
    CorrectedDataSheet['B24'] = f'MEAN+{howManySTD}STD (RawData)'
    CorrectedDataSheet['B24'].font = Font(sz=18)
    CorrectedDataSheet['B25'].font = Font(sz=18)


# Running Functions
RawDataSetup(rawDataSTD, rawDataMean)
CorrectedDataSetup(rawDataSTD, rawDataMean)


fileName = str(
    input('\nCompleted Successfully.\nPlease enter a name for the file: '))
wook.save(f'{workingPath}/{fileName}.xlsx')
